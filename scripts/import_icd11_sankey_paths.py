#!/usr/bin/env python3
"""Import ICD11 Sankey path rows from the WBE workbook into MySQL."""

from __future__ import annotations

import argparse
import os
import subprocess
import tempfile
from pathlib import Path
from typing import Iterable

import openpyxl


MAPPING_SHEET = "药物疾病ICD11映射"
IN_CHART_VALUE = "是"
INVALID_VALUES = {"", "NA", "N/A", "NULL", "NONE"}
REQUIRED_COLUMNS = [
    "目标类别",
    "目标物质类别",
    "目标物质子类",
    "药物",
    "适应症原文",
    "生物标记物名称",
    "biomarker",
    "规范适应症短语",
    "疾病实体短语",
    "ICD11_Level1_Code",
    "ICD11_Level1_Name",
    "ICD11_Level2_Code",
    "ICD11_Level2_Name",
    "ICD11_Level3_Code",
    "ICD11_Level3_Name",
    "映射层级",
    "匹配类型",
    "是否进入桑基图",
    "复核状态",
    "备注",
    "生物标记物CAS",
    "涉及文献数",
    "数据行数",
]
INSERT_COLUMNS = [
    "target_category",
    "substance_category",
    "substance_subclass",
    "drug_name",
    "indication_original",
    "biomarker_name",
    "biomarker_alias",
    "normalized_indication",
    "disease_entity",
    "icd11_level1_code",
    "icd11_level1_name",
    "icd11_level2_code",
    "icd11_level2_name",
    "icd11_level3_code",
    "icd11_level3_name",
    "mapping_level",
    "match_type",
    "review_status",
    "note",
    "biomarker_cas",
    "literature_count",
    "data_row_count",
]


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.upper() in INVALID_VALUES else text


def sql_value(value: object) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def decimal_text(value: object, fallback: str = "0") -> str:
    text = clean(value)
    if not text:
        return fallback
    try:
        return str(float(text))
    except ValueError:
        return fallback


def integer_text(value: object, fallback: str = "0") -> str:
    text = clean(value)
    if not text:
        return fallback
    try:
        return str(int(float(text)))
    except ValueError:
        return fallback


def required_value(row: tuple[object, ...], column: dict[str, int], name: str) -> str:
    return clean(row[column[name]])


def load_records(workbook_path: Path) -> tuple[list[tuple[object, ...]], dict[str, int]]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if MAPPING_SHEET not in workbook.sheetnames:
        raise ValueError(f"工作簿缺少工作表: {MAPPING_SHEET}")

    sheet = workbook[MAPPING_SHEET]
    header = [clean(value) for value in next(sheet.iter_rows(values_only=True))]
    column = {name: index for index, name in enumerate(header)}
    missing = [name for name in REQUIRED_COLUMNS if name not in column]
    if missing:
        raise ValueError(f"{MAPPING_SHEET} 缺少列: {', '.join(missing)}")

    records: list[tuple[object, ...]] = []
    stats = {"skipped": 0, "not_in_chart": 0}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if clean(row[column["是否进入桑基图"]]) != IN_CHART_VALUE:
            stats["not_in_chart"] += 1
            continue

        required_names = ["目标类别", "目标物质类别", "药物", "生物标记物名称", "ICD11_Level1_Name", "ICD11_Level2_Name"]
        if any(not required_value(row, column, name) for name in required_names):
            stats["skipped"] += 1
            continue

        records.append(
            (
                clean(row[column["目标类别"]]),
                clean(row[column["目标物质类别"]]),
                clean(row[column["目标物质子类"]]) or "未分类",
                clean(row[column["药物"]]),
                clean(row[column["适应症原文"]]),
                clean(row[column["生物标记物名称"]]),
                clean(row[column["biomarker"]]),
                clean(row[column["规范适应症短语"]]),
                clean(row[column["疾病实体短语"]]),
                clean(row[column["ICD11_Level1_Code"]]),
                clean(row[column["ICD11_Level1_Name"]]),
                clean(row[column["ICD11_Level2_Code"]]),
                clean(row[column["ICD11_Level2_Name"]]),
                clean(row[column["ICD11_Level3_Code"]]),
                clean(row[column["ICD11_Level3_Name"]]),
                clean(row[column["映射层级"]]),
                clean(row[column["匹配类型"]]),
                clean(row[column["复核状态"]]),
                clean(row[column["备注"]]),
                clean(row[column["生物标记物CAS"]]),
                decimal_text(row[column["涉及文献数"]], "1"),
                integer_text(row[column["数据行数"]], "0"),
            )
        )

    workbook.close()
    return records, stats


def write_import_sql(records: Iterable[tuple[object, ...]], output_path: Path, batch_size: int) -> None:
    record_list = list(records)
    with output_path.open("w", encoding="utf-8") as handle:
        handle.write("SET NAMES utf8mb4;\n")
        handle.write("TRUNCATE TABLE icd11_sankey_paths;\n")
        for start in range(0, len(record_list), batch_size):
            chunk = record_list[start : start + batch_size]
            handle.write(f"INSERT INTO icd11_sankey_paths ({', '.join(INSERT_COLUMNS)}) VALUES\n")
            values = []
            for record in chunk:
                values.append("(" + ", ".join(sql_value(value) for value in record) + ")")
            handle.write(",\n".join(values))
            handle.write(";\n")


def run_mysql(mysql_bin: str, database: str, user: str, password: str, sql_path: Path) -> None:
    command = [mysql_bin, f"-u{user}", database]
    env = os.environ.copy()
    if password:
        env["MYSQL_PWD"] = password
    with sql_path.open("rb") as sql_file:
        subprocess.run(command, stdin=sql_file, env=env, check=True)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    backend_root = Path(__file__).resolve().parents[1]
    workspace_root = backend_root.parent
    parser.add_argument("--workbook", default=str(workbook_path := workspace_root / "WBE汇总表6.25.xlsx"))
    parser.add_argument("--database", default="wbe")
    parser.add_argument("--user", default="root")
    parser.add_argument("--password", default="")
    parser.add_argument("--mysql-bin", default="mysql")
    parser.add_argument(
        "--schema",
        default=str(backend_root / "src/main/resources/db/icd11_sankey_schema.sql"),
    )
    parser.add_argument("--batch-size", type=int, default=300)
    args = parser.parse_args()

    source_workbook = Path(args.workbook).expanduser().resolve()
    records, stats = load_records(source_workbook)
    if not records:
        raise RuntimeError(f"没有可导入的 {MAPPING_SHEET} 入图记录: {workbook_path}")

    schema_path = Path(args.schema).expanduser().resolve()
    run_mysql(args.mysql_bin, args.database, args.user, args.password, schema_path)
    with tempfile.NamedTemporaryFile("w", suffix=".sql", delete=False, encoding="utf-8") as tmp:
        import_sql = Path(tmp.name)
    try:
        write_import_sql(records, import_sql, args.batch_size)
        run_mysql(args.mysql_bin, args.database, args.user, args.password, import_sql)
    finally:
        import_sql.unlink(missing_ok=True)

    print(
        f"Imported {len(records)} records into icd11_sankey_paths "
        f"(skipped={stats['skipped']}, not_in_chart={stats['not_in_chart']})."
    )


if __name__ == "__main__":
    main()

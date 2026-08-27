#!/usr/bin/env python3
"""Replace WBE business tables from a consolidated WBE workbook.

This script intentionally preserves system/reference tables such as users,
login logs, and geo_locations. It clears and reloads the workbook-backed
business tables.
"""

from __future__ import annotations

import argparse
import os
import re
import subprocess
import tempfile
import unicodedata
from collections import OrderedDict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

if __name__ == "__main__":
    from import_wbe_workbook_via_api import main as api_import_main

    api_import_main()
    raise SystemExit(0)

try:
    import openpyxl
except ModuleNotFoundError:
    openpyxl = None


DATA_SHEET = "数据表"
LITERATURE_SHEET = "文献基础信息"
ICD11_SHEET = "药物疾病ICD11映射"
INVALID_VALUES = {"", "NA", "N/A", "NULL", "NONE", "/"}
PRESCRIPTION_TYPES = {"处方药", "非处方药", "其他"}
ATC_TARGET_PREFIXES = set("ABCDGHJLMNPRSV")

REQUIRED_HEADERS = [
    "文献编号",
    "目标类别",
    "目标物质类别",
    "目标物质子类",
    "目标物质细类",
    "药物",
    "适应症",
    "处方/非处方",
    "生物标记物名称",
    "biomarker",
    "生物标记物CAS",
    "理化性质",
    "校准系数",
    "采样方法",
    "分析方法",
    "MDL_value",
    "MDL_unit",
    "MQL_value",
    "MQL_unit",
    "IDL_value",
    "IDL_unit",
    "IQL_value",
    "IQL_unit",
    "污水厂名称",
    "污水厂处理规模（m3/day）",
    "汇水区人群数量",
    "污水厂位置_国",
    "污水厂位置_省",
    "污水厂位置_市",
    "样品采集时间",
    "采样开始时间_YYYY_MM",
    "采样结束时间_YYYY_MM",
    "做图浓度_value",
    "做图浓度_unit",
    "进水浓度min_value",
    "进水浓度min_unit",
    "进水浓度max_value",
    "进水浓度max_unit",
    "进水浓度average_value",
    "进水浓度average_unit",
    "进水浓度median_value",
    "进水浓度median_unit",
    "每日质量负荷DLs",
    "DLs_unit",
    "PNDL_value",
    "PNDL_unit",
    "PNDL估算_value",
    "PNDL估算_unit",
    "做图PNDL_value",
    "做图PNDL_unit",
    "GS管道衰减系数",
    "人体排泄率（%）",
    "药物消费量_value",
    "药物消费量_unit",
    "药物使用流行率（%）",
    "疾病患病率（%）",
    "DOI",
    "keywords",
    "abstract",
]
OPTIONAL_HEADERS = ["来源工作簿说明", "原表行号说明"]

ICD11_REQUIRED_COLUMNS = [
    "目标类别",
    "目标物质类别",
    "目标物质子类",
    "目标物质细类",
    "药物",
    "适应症原文",
    "生物标记物名称",
    "biomarker",
    "规范适应症短语",
    "疾病实体短语",
    "ICD11_Level1_Code",
    "ICD11_Level1_Name",
    "ICD11_Level2_Code",
    "ICD11_Level2_Name",
    "ICD11_Level3_Code",
    "ICD11_Level3_Name",
    "映射层级",
    "匹配类型",
    "是否进入桑基图",
    "复核状态",
    "备注",
    "生物标记物CAS",
    "涉及文献数",
    "数据行数",
]

DECIMAL_FIELDS = [
    "校准系数",
    "MDL_value",
    "MQL_value",
    "IDL_value",
    "IQL_value",
    "做图浓度_value",
    "进水浓度min_value",
    "进水浓度max_value",
    "进水浓度average_value",
    "进水浓度median_value",
    "每日质量负荷DLs",
    "PNDL_value",
    "PNDL估算_value",
    "做图PNDL_value",
    "GS管道衰减系数",
    "人体排泄率（%）",
    "药物消费量_value",
    "药物使用流行率（%）",
    "疾病患病率（%）",
]

CREATE_SCHEMA_SQL = r"""
CREATE TABLE IF NOT EXISTS compounds (
    compound_id BIGINT PRIMARY KEY AUTO_INCREMENT COMMENT '化合物ID',
    target_category VARCHAR(160) NOT NULL COMMENT '目标类别',
    substance_category VARCHAR(180) NOT NULL COMMENT '目标物质类别',
    substance_subclass VARCHAR(180) DEFAULT NULL COMMENT '目标物质子类',
    substance_fine VARCHAR(180) DEFAULT NULL COMMENT '目标物质细类',
    drug_name VARCHAR(300) NOT NULL,
    indications LONGTEXT,
    prescription_type ENUM('处方药','非处方药','其他') DEFAULT NULL COMMENT '处方/非处方',
    biomarker_name VARCHAR(300) DEFAULT NULL,
    biomarker_cas VARCHAR(80) DEFAULT NULL COMMENT '生物标记物CAS',
    physicochemical_properties LONGTEXT,
    calibration_coefficient DECIMAL(15,6) DEFAULT NULL COMMENT '校准系数',
    human_excretion_rate DECIMAL(8,4) DEFAULT NULL COMMENT '人体排泄率(%)',
    consumption_value DECIMAL(20,4) DEFAULT NULL COMMENT '药物消费量_value',
    consumption_unit VARCHAR(80) DEFAULT NULL COMMENT '药物消费量_unit',
    usage_prevalence DECIMAL(8,4) DEFAULT NULL COMMENT '药物使用流行率(%)',
    disease_prevalence DECIMAL(8,4) DEFAULT NULL COMMENT '疾病患病率(%)',
    keywords LONGTEXT,
    doi VARCHAR(200) DEFAULT NULL,
    abstract LONGTEXT,
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
    UNIQUE KEY uk_drug_name (drug_name, biomarker_cas)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='化合物/药物信息表';

CREATE TABLE IF NOT EXISTS analytical_methods (
    method_id BIGINT PRIMARY KEY AUTO_INCREMENT,
    sampling_method VARCHAR(200) DEFAULT NULL COMMENT '采样方法',
    analysis_method VARCHAR(200) DEFAULT NULL COMMENT '分析方法',
    mdl_value DECIMAL(20,6) DEFAULT NULL COMMENT 'MDL_value',
    mdl_unit VARCHAR(30) DEFAULT NULL COMMENT 'MDL_unit',
    mql_value DECIMAL(20,6) DEFAULT NULL COMMENT 'MQL_value',
    mql_unit VARCHAR(30) DEFAULT NULL COMMENT 'MQL_unit',
    idl_value DECIMAL(20,6) DEFAULT NULL COMMENT 'IDL_value',
    idl_unit VARCHAR(30) DEFAULT NULL COMMENT 'IDL_unit',
    iql_value DECIMAL(20,6) DEFAULT NULL COMMENT 'IQL_value',
    iql_unit VARCHAR(30) DEFAULT NULL COMMENT 'IQL_unit',
    UNIQUE KEY uk_method (sampling_method, analysis_method)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='采样与分析方法表';

CREATE TABLE IF NOT EXISTS wastewater_plants (
    plant_id BIGINT PRIMARY KEY AUTO_INCREMENT,
    plant_name VARCHAR(500) NOT NULL COMMENT '污水厂名称',
    treatment_capacity_m3_day BIGINT DEFAULT NULL COMMENT '处理规模(m3/day)',
    served_population BIGINT DEFAULT NULL COMMENT '汇水区人群数量',
    country VARCHAR(100) DEFAULT NULL COMMENT '国家',
    province VARCHAR(100) DEFAULT NULL COMMENT '省份',
    city VARCHAR(100) DEFAULT NULL COMMENT '城市',
    gs_attenuation_coefficient DECIMAL(10,6) DEFAULT NULL COMMENT 'GS管道衰减系数',
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    UNIQUE KEY uk_plant_name (plant_name)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='污水厂信息表';

CREATE TABLE IF NOT EXISTS sampling_events (
    event_id BIGINT PRIMARY KEY AUTO_INCREMENT,
    plant_id BIGINT NOT NULL COMMENT '污水厂ID',
    sample_collection_time DATETIME DEFAULT NULL COMMENT '样品采集时间',
    sampling_start_ym VARCHAR(7) DEFAULT NULL COMMENT '采样开始时间_YYYY_MM',
    sampling_end_ym VARCHAR(7) DEFAULT NULL COMMENT '采样结束时间_YYYY_MM',
    source_workbook VARCHAR(255) DEFAULT NULL COMMENT '来源工作簿说明',
    original_row_number INT DEFAULT NULL COMMENT '原表行号',
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    KEY plant_id (plant_id),
    KEY idx_sampling_time (sampling_start_ym, sampling_end_ym),
    CONSTRAINT sampling_events_ibfk_1 FOREIGN KEY (plant_id) REFERENCES wastewater_plants (plant_id) ON DELETE RESTRICT
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='采样事件表';

CREATE TABLE IF NOT EXISTS measurements (
    measurement_id BIGINT PRIMARY KEY AUTO_INCREMENT,
    compound_id BIGINT NOT NULL COMMENT '化合物ID',
    method_id BIGINT NOT NULL COMMENT '分析方法ID',
    event_id BIGINT NOT NULL COMMENT '采样事件ID',
    plot_concentration_value DECIMAL(20,6) DEFAULT NULL COMMENT '做图浓度_value',
    plot_concentration_unit VARCHAR(30) DEFAULT NULL COMMENT '做图浓度_unit',
    inflow_min_value DECIMAL(20,6) DEFAULT NULL COMMENT '进水浓度min_value',
    inflow_min_unit VARCHAR(30) DEFAULT NULL COMMENT '进水浓度min_unit',
    inflow_max_value DECIMAL(20,6) DEFAULT NULL COMMENT '进水浓度max_value',
    inflow_max_unit VARCHAR(30) DEFAULT NULL COMMENT '进水浓度max_unit',
    inflow_avg_value DECIMAL(20,6) DEFAULT NULL COMMENT '进水浓度average_value',
    inflow_avg_unit VARCHAR(30) DEFAULT NULL COMMENT '进水浓度average_unit',
    inflow_median_value DECIMAL(20,6) DEFAULT NULL COMMENT '进水浓度median_value',
    inflow_median_unit VARCHAR(30) DEFAULT NULL COMMENT '进水浓度median_unit',
    daily_load_dls_value DECIMAL(25,6) DEFAULT NULL COMMENT '每日质量负荷DLs',
    daily_load_dls_unit VARCHAR(30) DEFAULT NULL COMMENT 'DLs_unit',
    pndl_value DECIMAL(25,6) DEFAULT NULL COMMENT 'PNDL_value',
    pndl_unit VARCHAR(30) DEFAULT NULL COMMENT 'PNDL_unit',
    pndl_estimated_value DECIMAL(25,6) DEFAULT NULL COMMENT 'PNDL估算_value',
    pndl_estimated_unit VARCHAR(30) DEFAULT NULL COMMENT 'PNDL估算_unit',
    plot_pndl_value DECIMAL(25,6) DEFAULT NULL COMMENT '做图PNDL_value',
    plot_pndl_unit VARCHAR(30) DEFAULT NULL COMMENT '做图PNDL_unit',
    KEY method_id (method_id),
    KEY event_id (event_id),
    KEY idx_compound_event (compound_id, event_id),
    CONSTRAINT measurements_ibfk_1 FOREIGN KEY (compound_id) REFERENCES compounds (compound_id) ON DELETE RESTRICT,
    CONSTRAINT measurements_ibfk_2 FOREIGN KEY (method_id) REFERENCES analytical_methods (method_id) ON DELETE RESTRICT,
    CONSTRAINT measurements_ibfk_3 FOREIGN KEY (event_id) REFERENCES sampling_events (event_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='测量结果事实表';

CREATE TABLE IF NOT EXISTS home_target_records (
    record_id BIGINT PRIMARY KEY AUTO_INCREMENT,
    literature_id VARCHAR(50) NOT NULL COMMENT '文献编号',
    doi VARCHAR(200) NOT NULL COMMENT 'DOI',
    target_category VARCHAR(160) NOT NULL COMMENT '目标类别',
    target_group VARCHAR(20) NOT NULL COMMENT '首页目标组: drug/consumer',
    substance_category VARCHAR(180) NOT NULL COMMENT '目标物质类别',
    substance_subclass VARCHAR(180) NOT NULL COMMENT '目标物质子类',
    substance_fine VARCHAR(180) DEFAULT NULL COMMENT '目标物质细类',
    biomarker_name VARCHAR(300) NOT NULL COMMENT '生物标记物名称',
    source_sheet VARCHAR(64) NOT NULL DEFAULT '数据表',
    source_row_number INT NOT NULL COMMENT 'Excel 原始行号',
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    UNIQUE KEY uk_home_target_source_row (source_sheet, source_row_number),
    INDEX idx_home_target_group_category (target_group, substance_category),
    INDEX idx_home_target_category (substance_category),
    INDEX idx_home_target_subclass (substance_category, substance_subclass),
    INDEX idx_home_target_biomarker (substance_category, substance_subclass, biomarker_name),
    INDEX idx_home_target_doi (doi)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='首页目标物质研究图逐行事实表';

CREATE TABLE IF NOT EXISTS icd11_sankey_paths (
    sankey_path_id BIGINT PRIMARY KEY AUTO_INCREMENT,
    target_category VARCHAR(160) NOT NULL COMMENT '目标类别',
    substance_category VARCHAR(180) NOT NULL COMMENT '目标物质类别',
    substance_subclass VARCHAR(180) NOT NULL COMMENT '目标物质子类',
    substance_fine VARCHAR(180) DEFAULT NULL COMMENT '目标物质细类',
    drug_name VARCHAR(300) NOT NULL COMMENT '药物名称',
    indication_original TEXT COMMENT '适应症原文',
    biomarker_name VARCHAR(300) NOT NULL COMMENT '生物标记物名称',
    biomarker_alias VARCHAR(300) COMMENT '英文 biomarker 或别名',
    normalized_indication VARCHAR(300) COMMENT '规范适应症短语',
    disease_entity VARCHAR(300) COMMENT '疾病实体短语',
    icd11_level1_code VARCHAR(80),
    icd11_level1_name VARCHAR(220) NOT NULL,
    icd11_level2_code VARCHAR(80),
    icd11_level2_name VARCHAR(220) NOT NULL,
    icd11_level3_code VARCHAR(80),
    icd11_level3_name VARCHAR(220),
    mapping_level VARCHAR(80),
    match_type VARCHAR(180),
    review_status VARCHAR(120),
    note TEXT,
    biomarker_cas VARCHAR(80),
    literature_count DECIMAL(18,4) NOT NULL DEFAULT 1 COMMENT '桑基图权重：涉及文献数',
    data_row_count BIGINT NOT NULL DEFAULT 0 COMMENT '后台统计：数据行数',
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    KEY idx_icd11_sankey_category (target_category),
    KEY idx_icd11_sankey_level1 (target_category, icd11_level1_name),
    KEY idx_icd11_sankey_level2 (target_category, icd11_level2_name),
    KEY idx_icd11_sankey_drug (target_category, drug_name),
    KEY idx_icd11_sankey_biomarker (target_category, biomarker_name)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='ICD11 四层桑基图聚合路径表';

CREATE TABLE IF NOT EXISTS geo_locations (
    geo_location_id BIGINT PRIMARY KEY AUTO_INCREMENT,
    level VARCHAR(16) NOT NULL,
    geo_key VARCHAR(180) NOT NULL,
    parent_geo_key VARCHAR(180),
    country VARCHAR(120),
    province VARCHAR(120),
    city VARCHAR(120),
    display_name VARCHAR(180) NOT NULL,
    latitude DECIMAL(12,7),
    longitude DECIMAL(12,7),
    is_mappable BOOLEAN NOT NULL DEFAULT TRUE,
    coordinate_source VARCHAR(80),
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
    UNIQUE KEY uk_geo_level_key (level, geo_key),
    KEY idx_geo_parent (level, parent_geo_key),
    KEY idx_geo_names (country, province, city)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='地图可视化地理维表';

CREATE TABLE IF NOT EXISTS map_pndl_stats (
    stat_id BIGINT PRIMARY KEY AUTO_INCREMENT,
    level VARCHAR(16) NOT NULL,
    geo_key VARCHAR(180) NOT NULL,
    parent_geo_key VARCHAR(180),
    display_name VARCHAR(180) NOT NULL,
    country VARCHAR(120),
    province VARCHAR(120),
    city VARCHAR(120),
    latitude DECIMAL(12,7),
    longitude DECIMAL(12,7),
    target_class VARCHAR(120),
    category VARCHAR(120) NOT NULL,
    subcategory VARCHAR(120) NOT NULL,
    biomarker_key VARCHAR(80) NOT NULL,
    biomarker_label VARCHAR(220) NOT NULL,
    biomarker_cas VARCHAR(80),
    year_label VARCHAR(32) NOT NULL,
    pndl_geomean_mg_d_1000inh DECIMAL(28,10),
    pndl_mean_mg_d_1000inh DECIMAL(28,10),
    pndl_min_mg_d_1000inh DECIMAL(28,10),
    pndl_max_mg_d_1000inh DECIMAL(28,10),
    record_count BIGINT NOT NULL DEFAULT 0,
    doi_count BIGINT NOT NULL DEFAULT 0,
    year_count BIGINT NOT NULL DEFAULT 0,
    city_count BIGINT NOT NULL DEFAULT 0,
    point_count BIGINT NOT NULL DEFAULT 0,
    pndl_sources VARCHAR(200),
    is_mappable BOOLEAN NOT NULL DEFAULT TRUE,
    refreshed_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    KEY idx_map_filter (category, subcategory, biomarker_key, year_label, level, geo_key),
    KEY idx_map_geo (level, geo_key),
    KEY idx_map_value (pndl_geomean_mg_d_1000inh)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='地图可视化PNDL预聚合表';

ALTER TABLE compounds ADD COLUMN IF NOT EXISTS substance_fine VARCHAR(180) DEFAULT NULL COMMENT '目标物质细类';
ALTER TABLE home_target_records ADD COLUMN IF NOT EXISTS substance_fine VARCHAR(180) DEFAULT NULL COMMENT '目标物质细类';
ALTER TABLE icd11_sankey_paths ADD COLUMN IF NOT EXISTS substance_fine VARCHAR(180) DEFAULT NULL COMMENT '目标物质细类';
ALTER TABLE compounds
    MODIFY target_category VARCHAR(160) NOT NULL COMMENT '目标类别',
    MODIFY substance_category VARCHAR(180) NOT NULL COMMENT '目标物质类别',
    MODIFY substance_subclass VARCHAR(180) DEFAULT NULL COMMENT '目标物质子类',
    MODIFY biomarker_cas VARCHAR(80) DEFAULT NULL COMMENT '生物标记物CAS',
    MODIFY consumption_unit VARCHAR(80) DEFAULT NULL COMMENT '药物消费量_unit';
ALTER TABLE wastewater_plants MODIFY plant_name VARCHAR(500) NOT NULL COMMENT '污水厂名称';
ALTER TABLE home_target_records
    MODIFY target_category VARCHAR(160) NOT NULL COMMENT '目标类别',
    MODIFY substance_category VARCHAR(180) NOT NULL COMMENT '目标物质类别',
    MODIFY substance_subclass VARCHAR(180) NOT NULL COMMENT '目标物质子类';
"""


@dataclass(frozen=True)
class WorkbookData:
    data_rows: list[dict[str, str]]
    compounds: OrderedDict[tuple[str, str | None], dict[str, str]]
    methods: OrderedDict[tuple[str, str], dict[str, str]]
    plants: OrderedDict[str, dict[str, str]]
    literature_dois: dict[str, str]
    icd11_rows: list[dict[str, str]]
    stats: dict[str, int]


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.upper() in INVALID_VALUES else text


def raw_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def is_useful(value: str | None) -> bool:
    return clean(value) != ""


def sql_string(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def sql_decimal(value: Decimal | None) -> str:
    if value is None:
        return "NULL"
    return format(value, "f")


def sql_integer(value: int | None) -> str:
    return "NULL" if value is None else str(value)


def parse_decimal(value: str | None) -> Decimal | None:
    text = raw_text(value)
    if not is_useful(text):
        return None
    compact = text.replace(",", "").replace("，", "")
    try:
        return Decimal(compact)
    except InvalidOperation:
        pass
    upper = compact.upper()
    if compact.startswith("<") or compact.startswith(">") or upper in {"ND", "BQL"}:
        return None
    preferred = re.search(r"取\s*([+-]?\d+(?:\.\d+)?)", compact)
    if preferred:
        return Decimal(preferred.group(1))
    if "~" in compact or "～" in compact or re.search(r"\d\s*-\s*\d", compact):
        return None
    leading = re.match(r"^([+-]?\d+(?:\.\d+)?)", compact)
    if leading:
        return Decimal(leading.group(1))
    return None


def parse_int(value: str | None) -> int | None:
    decimal = parse_decimal(value)
    if decimal is None:
        return None
    return int(decimal.to_integral_value())


def decimal_text(value: str | None, fallback: str = "0") -> str:
    decimal = parse_decimal(value)
    return fallback if decimal is None else format(decimal, "f")


def target_group(target_category: str) -> str:
    normalized = clean(target_category)
    if normalized == "药物类":
        return "drug"
    if normalized[:1] in ATC_TARGET_PREFIXES and (len(normalized) == 1 or normalized[1].isspace()):
        return "drug"
    if "药物" in normalized and not any(token in normalized for token in ("消费", "生活方式", "暴露")):
        return "drug"
    return "consumer"


def load_workbook_data(workbook_path: Path) -> WorkbookData:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if DATA_SHEET not in workbook.sheetnames:
            raise ValueError(f"工作簿缺少工作表: {DATA_SHEET}")
        data_sheet = workbook[DATA_SHEET]
        header = [clean(value) for value in next(data_sheet.iter_rows(values_only=True))]
        if header[: len(REQUIRED_HEADERS)] != REQUIRED_HEADERS:
            for index, expected in enumerate(REQUIRED_HEADERS):
                actual = header[index] if index < len(header) else ""
                if actual != expected:
                    raise ValueError(f"{DATA_SHEET} 第 {index + 1} 列应为“{expected}”，实际为“{actual or '空'}”")
        all_headers = REQUIRED_HEADERS + OPTIONAL_HEADERS
        data_rows: list[dict[str, str]] = []
        compounds: OrderedDict[tuple[str, str | None], dict[str, str]] = OrderedDict()
        methods: OrderedDict[tuple[str, str], dict[str, str]] = OrderedDict()
        plants: OrderedDict[str, dict[str, str]] = OrderedDict()
        stats = {
            "data_rows": 0,
            "skipped_blank": 0,
            "validation_errors": 0,
            "home_records": 0,
            "home_missing_doi": 0,
            "icd11_records": 0,
            "icd11_not_in_chart": 0,
            "icd11_skipped": 0,
        }
        validation_examples: list[str] = []

        for excel_row_number, row in enumerate(data_sheet.iter_rows(min_row=2, values_only=True), start=2):
            values = {name: clean(row[index] if index < len(row) else None) for index, name in enumerate(all_headers)}
            if all(not values.get(header_name) for header_name in REQUIRED_HEADERS):
                stats["skipped_blank"] += 1
                continue
            errors = validate_data_row(values)
            if errors:
                stats["validation_errors"] += 1
                if len(validation_examples) < 10:
                    validation_examples.append(f"第 {excel_row_number} 行: {'; '.join(errors)}")
                continue

            stats["data_rows"] += 1
            data_rows.append(values)

            compounds.setdefault(compound_unique_key(values), values)
            methods.setdefault(method_unique_key(values), values)
            plants.setdefault(plant_unique_key(values), values)

        if validation_examples:
            raise ValueError("数据表存在阻断错误，已停止导入:\n" + "\n".join(validation_examples))

        literature_dois = load_literature_dois(workbook)
        icd11_rows = load_icd11_rows(workbook, stats)

        return WorkbookData(data_rows, compounds, methods, plants, literature_dois, icd11_rows, stats)
    finally:
        workbook.close()


def validate_data_row(values: dict[str, str]) -> list[str]:
    errors: list[str] = []
    for required in ["文献编号", "目标类别", "目标物质类别", "采样方法", "分析方法"]:
        if not is_useful(values.get(required)):
            errors.append(f"{required} 不能为空或 NA")
    if not any(is_useful(values.get(name)) for name in ["药物", "生物标记物名称", "biomarker"]):
        errors.append("药物、生物标记物名称、biomarker 至少需要填写一个")
    return errors


def first_useful(*values: str) -> str:
    for value in values:
        cleaned = clean(value)
        if cleaned:
            return cleaned
    return "NA"


def mysql_text_key(value: str | None) -> str:
    """Approximate utf8mb4_0900_ai_ci comparisons for import-side de-duping."""
    text = clean(value).casefold()
    decomposed = unicodedata.normalize("NFKD", text)
    return "".join(char for char in decomposed if not unicodedata.combining(char))


def compound_unique_key(data: dict[str, str]) -> tuple[str, str | None]:
    drug_name = first_useful(data.get("药物"), data.get("生物标记物名称"), data.get("biomarker"), "NA")
    biomarker_cas = clean(data.get("生物标记物CAS")) or None
    return mysql_text_key(drug_name), mysql_text_key(biomarker_cas) if biomarker_cas is not None else None


def method_unique_key(data: dict[str, str]) -> tuple[str, str]:
    return mysql_text_key(first_useful(data.get("采样方法"), "NA")), mysql_text_key(first_useful(data.get("分析方法"), "NA"))


def plant_unique_key(data: dict[str, str]) -> str:
    return mysql_text_key(first_useful(data.get("污水厂名称"), "NA"))


def load_literature_dois(workbook: openpyxl.Workbook) -> dict[str, str]:
    if LITERATURE_SHEET not in workbook.sheetnames:
        return {}
    sheet = workbook[LITERATURE_SHEET]
    header = [clean(value) for value in next(sheet.iter_rows(values_only=True))]
    column = {name: index for index, name in enumerate(header)}
    if "文献编号" not in column or "DOI" not in column:
        raise ValueError(f"{LITERATURE_SHEET} 缺少列: 文献编号, DOI")
    dois: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        literature_id = clean(row[column["文献编号"]])
        doi = clean(row[column["DOI"]])
        if literature_id and doi:
            dois[literature_id] = doi
    return dois


def load_icd11_rows(workbook: openpyxl.Workbook, stats: dict[str, int]) -> list[dict[str, str]]:
    if ICD11_SHEET not in workbook.sheetnames:
        return []
    sheet = workbook[ICD11_SHEET]
    header = [clean(value) for value in next(sheet.iter_rows(values_only=True))]
    column = {name: index for index, name in enumerate(header)}
    missing = [name for name in ICD11_REQUIRED_COLUMNS if name not in column]
    if missing:
        raise ValueError(f"{ICD11_SHEET} 缺少列: {', '.join(missing)}")

    rows: list[dict[str, str]] = []
    required_names = ["目标类别", "目标物质类别", "药物", "生物标记物名称", "ICD11_Level1_Name", "ICD11_Level2_Name"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = {name: clean(row[column[name]]) for name in ICD11_REQUIRED_COLUMNS}
        if item["是否进入桑基图"] != "是":
            stats["icd11_not_in_chart"] += 1
            continue
        if any(not item[name] for name in required_names):
            stats["icd11_skipped"] += 1
            continue
        rows.append(item)
    stats["icd11_records"] = len(rows)
    return rows


def write_sql(
    data: WorkbookData,
    output_path: Path,
    map_refresh_path: Path,
    batch_size: int,
    clear_upload_audit: bool,
    existing_tables: set[str],
) -> None:
    with output_path.open("w", encoding="utf-8") as handle:
        handle.write("SET NAMES utf8mb4;\n")
        handle.write(CREATE_SCHEMA_SQL)
        handle.write("\nSET FOREIGN_KEY_CHECKS=0;\n")
        for table in [
            "map_pndl_stats",
            "measurements",
            "sampling_events",
            "analytical_methods",
            "wastewater_plants",
            "compounds",
            "home_target_records",
            "icd11_sankey_paths",
        ]:
            handle.write(f"TRUNCATE TABLE {table};\n")
        if clear_upload_audit:
            for table in ["data_upload_rows", "data_upload_batches"]:
                if table in existing_tables:
                    handle.write(f"TRUNCATE TABLE {table};\n")
        handle.write("SET FOREIGN_KEY_CHECKS=1;\n\n")

        write_compounds(handle, data.compounds.values(), batch_size)
        write_methods(handle, data.methods.values(), batch_size)
        write_plants(handle, data.plants.values(), batch_size)
        write_sampling_events(handle, data.data_rows, data.plants, batch_size)
        write_measurements(handle, data.data_rows, data.compounds, data.methods, batch_size)
        write_home_records(handle, data, batch_size)
        write_icd11_paths(handle, data.icd11_rows, batch_size)

        if map_refresh_path.exists():
            handle.write("\n")
            handle.write(map_refresh_path.read_text(encoding="utf-8"))
            handle.write("\n")


def write_insert(
    handle,
    table: str,
    columns: list[str],
    rows: Iterable[list[str]],
    batch_size: int,
) -> int:
    rows_written = 0
    batch: list[list[str]] = []
    for row in rows:
        batch.append(row)
        if len(batch) >= batch_size:
            flush_insert(handle, table, columns, batch)
            rows_written += len(batch)
            batch.clear()
    if batch:
        flush_insert(handle, table, columns, batch)
        rows_written += len(batch)
    return rows_written


def flush_insert(handle, table: str, columns: list[str], rows: list[list[str]]) -> None:
    handle.write(f"INSERT INTO {table} ({', '.join(columns)}) VALUES\n")
    handle.write(",\n".join("(" + ", ".join(row) + ")" for row in rows))
    handle.write(";\n")


def write_compounds(handle, rows: Iterable[dict[str, str]], batch_size: int) -> None:
    columns = [
        "compound_id",
        "target_category",
        "substance_category",
        "substance_subclass",
        "substance_fine",
        "drug_name",
        "indications",
        "prescription_type",
        "biomarker_name",
        "biomarker_cas",
        "physicochemical_properties",
        "calibration_coefficient",
        "human_excretion_rate",
        "consumption_value",
        "consumption_unit",
        "usage_prevalence",
        "disease_prevalence",
        "keywords",
        "doi",
        "abstract",
    ]
    sql_rows = []
    for compound_id, data in enumerate(rows, start=1):
        drug_name = first_useful(data.get("药物"), data.get("生物标记物名称"), data.get("biomarker"), "NA")
        prescription = clean(data.get("处方/非处方"))
        sql_rows.append(
            [
                str(compound_id),
                sql_string(first_useful(data.get("目标类别"), "NA")),
                sql_string(first_useful(data.get("目标物质类别"), "NA")),
                sql_string(clean(data.get("目标物质子类")) or None),
                sql_string(clean(data.get("目标物质细类")) or None),
                sql_string(drug_name),
                sql_string(clean(data.get("适应症")) or None),
                sql_string(prescription if prescription in PRESCRIPTION_TYPES else None),
                sql_string(clean(data.get("生物标记物名称")) or None),
                sql_string(clean(data.get("生物标记物CAS")) or None),
                sql_string(clean(data.get("理化性质")) or None),
                sql_decimal(parse_decimal(data.get("校准系数"))),
                sql_decimal(parse_decimal(data.get("人体排泄率（%）"))),
                sql_decimal(parse_decimal(data.get("药物消费量_value"))),
                sql_string(clean(data.get("药物消费量_unit")) or None),
                sql_decimal(parse_decimal(data.get("药物使用流行率（%）"))),
                sql_decimal(parse_decimal(data.get("疾病患病率（%）"))),
                sql_string(clean(data.get("keywords")) or None),
                sql_string(clean(data.get("DOI")) or None),
                sql_string(clean(data.get("abstract")) or None),
            ]
        )
    write_insert(handle, "compounds", columns, sql_rows, batch_size)


def write_methods(handle, rows: Iterable[dict[str, str]], batch_size: int) -> None:
    columns = [
        "method_id",
        "sampling_method",
        "analysis_method",
        "mdl_value",
        "mdl_unit",
        "mql_value",
        "mql_unit",
        "idl_value",
        "idl_unit",
        "iql_value",
        "iql_unit",
    ]
    sql_rows = []
    for method_id, data in enumerate(rows, start=1):
        sql_rows.append(
            [
                str(method_id),
                sql_string(first_useful(data.get("采样方法"), "NA")),
                sql_string(first_useful(data.get("分析方法"), "NA")),
                sql_decimal(parse_decimal(data.get("MDL_value"))),
                sql_string(clean(data.get("MDL_unit")) or None),
                sql_decimal(parse_decimal(data.get("MQL_value"))),
                sql_string(clean(data.get("MQL_unit")) or None),
                sql_decimal(parse_decimal(data.get("IDL_value"))),
                sql_string(clean(data.get("IDL_unit")) or None),
                sql_decimal(parse_decimal(data.get("IQL_value"))),
                sql_string(clean(data.get("IQL_unit")) or None),
            ]
        )
    write_insert(handle, "analytical_methods", columns, sql_rows, batch_size)


def write_plants(handle, rows: Iterable[dict[str, str]], batch_size: int) -> None:
    columns = [
        "plant_id",
        "plant_name",
        "treatment_capacity_m3_day",
        "served_population",
        "country",
        "province",
        "city",
        "gs_attenuation_coefficient",
    ]
    sql_rows = []
    for plant_id, data in enumerate(rows, start=1):
        sql_rows.append(
            [
                str(plant_id),
                sql_string(first_useful(data.get("污水厂名称"), "NA")),
                sql_integer(parse_int(data.get("污水厂处理规模（m3/day）"))),
                sql_integer(parse_int(data.get("汇水区人群数量"))),
                sql_string(clean(data.get("污水厂位置_国")) or None),
                sql_string(clean(data.get("污水厂位置_省")) or None),
                sql_string(clean(data.get("污水厂位置_市")) or None),
                sql_decimal(parse_decimal(data.get("GS管道衰减系数"))),
            ]
        )
    write_insert(handle, "wastewater_plants", columns, sql_rows, batch_size)


def write_sampling_events(
    handle,
    rows: list[dict[str, str]],
    plants: OrderedDict[str, dict[str, str]],
    batch_size: int,
) -> None:
    plant_ids = {plant_key: index for index, plant_key in enumerate(plants.keys(), start=1)}
    columns = [
        "event_id",
        "plant_id",
        "sample_collection_time",
        "sampling_start_ym",
        "sampling_end_ym",
        "source_workbook",
        "original_row_number",
    ]
    sql_rows = []
    for event_id, data in enumerate(rows, start=1):
        plant_key = plant_unique_key(data)
        sql_rows.append(
            [
                str(event_id),
                str(plant_ids[plant_key]),
                "NULL",
                sql_string(clean(data.get("采样开始时间_YYYY_MM")) or None),
                sql_string(clean(data.get("采样结束时间_YYYY_MM")) or None),
                sql_string(clean(data.get("来源工作簿说明")) or None),
                sql_integer(parse_int(data.get("原表行号说明"))),
            ]
        )
    write_insert(handle, "sampling_events", columns, sql_rows, batch_size)


def write_measurements(
    handle,
    rows: list[dict[str, str]],
    compounds: OrderedDict[tuple[str, str | None], dict[str, str]],
    methods: OrderedDict[tuple[str, str], dict[str, str]],
    batch_size: int,
) -> None:
    compound_ids = {key: index for index, key in enumerate(compounds.keys(), start=1)}
    method_ids = {key: index for index, key in enumerate(methods.keys(), start=1)}
    columns = [
        "measurement_id",
        "compound_id",
        "method_id",
        "event_id",
        "plot_concentration_value",
        "plot_concentration_unit",
        "inflow_min_value",
        "inflow_min_unit",
        "inflow_max_value",
        "inflow_max_unit",
        "inflow_avg_value",
        "inflow_avg_unit",
        "inflow_median_value",
        "inflow_median_unit",
        "daily_load_dls_value",
        "daily_load_dls_unit",
        "pndl_value",
        "pndl_unit",
        "pndl_estimated_value",
        "pndl_estimated_unit",
        "plot_pndl_value",
        "plot_pndl_unit",
    ]
    sql_rows = []
    for measurement_id, data in enumerate(rows, start=1):
        compound_key = compound_unique_key(data)
        method_key = method_unique_key(data)
        sql_rows.append(
            [
                str(measurement_id),
                str(compound_ids[compound_key]),
                str(method_ids[method_key]),
                str(measurement_id),
                sql_decimal(parse_decimal(data.get("做图浓度_value"))),
                sql_string(clean(data.get("做图浓度_unit")) or None),
                sql_decimal(parse_decimal(data.get("进水浓度min_value"))),
                sql_string(clean(data.get("进水浓度min_unit")) or None),
                sql_decimal(parse_decimal(data.get("进水浓度max_value"))),
                sql_string(clean(data.get("进水浓度max_unit")) or None),
                sql_decimal(parse_decimal(data.get("进水浓度average_value"))),
                sql_string(clean(data.get("进水浓度average_unit")) or None),
                sql_decimal(parse_decimal(data.get("进水浓度median_value"))),
                sql_string(clean(data.get("进水浓度median_unit")) or None),
                sql_decimal(parse_decimal(data.get("每日质量负荷DLs"))),
                sql_string(clean(data.get("DLs_unit")) or None),
                sql_decimal(parse_decimal(data.get("PNDL_value"))),
                sql_string(clean(data.get("PNDL_unit")) or None),
                sql_decimal(parse_decimal(data.get("PNDL估算_value"))),
                sql_string(clean(data.get("PNDL估算_unit")) or None),
                sql_decimal(parse_decimal(data.get("做图PNDL_value"))),
                sql_string(clean(data.get("做图PNDL_unit")) or None),
            ]
        )
    write_insert(handle, "measurements", columns, sql_rows, batch_size)


def write_home_records(handle, data: WorkbookData, batch_size: int) -> None:
    columns = [
        "literature_id",
        "doi",
        "target_category",
        "target_group",
        "substance_category",
        "substance_subclass",
        "substance_fine",
        "biomarker_name",
        "source_sheet",
        "source_row_number",
    ]
    sql_rows = []
    for row_number, row in enumerate(data.data_rows, start=2):
        literature_id = clean(row.get("文献编号"))
        target_category = clean(row.get("目标类别"))
        substance_category = clean(row.get("目标物质类别"))
        substance_subclass = clean(row.get("目标物质子类")) or "默认"
        biomarker_name = clean(row.get("生物标记物名称"))
        doi = data.literature_dois.get(literature_id) or clean(row.get("DOI"))
        if not doi:
            data.stats["home_missing_doi"] += 1
            continue
        if not all([literature_id, target_category, substance_category, biomarker_name]):
            continue
        sql_rows.append(
            [
                sql_string(literature_id),
                sql_string(doi),
                sql_string(target_category),
                sql_string(target_group(target_category)),
                sql_string(substance_category),
                sql_string(substance_subclass),
                sql_string(clean(row.get("目标物质细类")) or None),
                sql_string(biomarker_name),
                sql_string(DATA_SHEET),
                str(row_number),
            ]
        )
    data.stats["home_records"] = len(sql_rows)
    write_insert(handle, "home_target_records", columns, sql_rows, batch_size)


def write_icd11_paths(handle, rows: list[dict[str, str]], batch_size: int) -> None:
    columns = [
        "target_category",
        "substance_category",
        "substance_subclass",
        "substance_fine",
        "drug_name",
        "indication_original",
        "biomarker_name",
        "biomarker_alias",
        "normalized_indication",
        "disease_entity",
        "icd11_level1_code",
        "icd11_level1_name",
        "icd11_level2_code",
        "icd11_level2_name",
        "icd11_level3_code",
        "icd11_level3_name",
        "mapping_level",
        "match_type",
        "review_status",
        "note",
        "biomarker_cas",
        "literature_count",
        "data_row_count",
    ]
    sql_rows = []
    for row in rows:
        sql_rows.append(
            [
                sql_string(row["目标类别"]),
                sql_string(row["目标物质类别"]),
                sql_string(row["目标物质子类"] or "未分类"),
                sql_string(row["目标物质细类"] or None),
                sql_string(row["药物"]),
                sql_string(row["适应症原文"] or None),
                sql_string(row["生物标记物名称"]),
                sql_string(row["biomarker"] or None),
                sql_string(row["规范适应症短语"] or None),
                sql_string(row["疾病实体短语"] or None),
                sql_string(row["ICD11_Level1_Code"] or None),
                sql_string(row["ICD11_Level1_Name"]),
                sql_string(row["ICD11_Level2_Code"] or None),
                sql_string(row["ICD11_Level2_Name"]),
                sql_string(row["ICD11_Level3_Code"] or None),
                sql_string(row["ICD11_Level3_Name"] or None),
                sql_string(row["映射层级"] or None),
                sql_string(row["匹配类型"] or None),
                sql_string(row["复核状态"] or None),
                sql_string(row["备注"] or None),
                sql_string(row["生物标记物CAS"] or None),
                decimal_text(row["涉及文献数"], "1"),
                str(parse_int(row["数据行数"]) or 0),
            ]
        )
    write_insert(handle, "icd11_sankey_paths", columns, sql_rows, batch_size)


def mysql_command(mysql_bin: str, database: str, user: str, password: str) -> list[str]:
    return [mysql_bin, f"-u{user}", database]


def mysql_env(password: str) -> dict[str, str]:
    env = os.environ.copy()
    if password:
        env["MYSQL_PWD"] = password
    return env


def load_existing_tables(mysql_bin: str, database: str, user: str, password: str) -> set[str]:
    result = subprocess.run(
        mysql_command(mysql_bin, database, user, password) + ["-N", "-B", "-e", "SHOW TABLES"],
        env=mysql_env(password),
        check=True,
        text=True,
        stdout=subprocess.PIPE,
    )
    return {line.strip() for line in result.stdout.splitlines() if line.strip()}


def run_mysql(mysql_bin: str, database: str, user: str, password: str, sql_path: Path) -> None:
    with sql_path.open("rb") as sql_file:
        subprocess.run(mysql_command(mysql_bin, database, user, password), stdin=sql_file, env=mysql_env(password), check=True)


def main() -> None:
    backend_root = Path(__file__).resolve().parents[1]
    workspace_root = backend_root.parent
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", default=str(workspace_root / "WBE汇总表6.29.xlsx"))
    parser.add_argument("--database", default="wbe")
    parser.add_argument("--user", default="root")
    parser.add_argument("--password", default="")
    parser.add_argument("--mysql-bin", default="mysql")
    parser.add_argument("--batch-size", type=int, default=300)
    parser.add_argument("--dry-run", action="store_true", help="Parse and generate SQL without applying it.")
    parser.add_argument(
        "--preserve-upload-audit",
        action="store_true",
        help="Keep data_upload_batches/data_upload_rows instead of clearing upload history.",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    map_refresh_path = backend_root / "src/main/resources/db/map_pndl_stats_refresh.sql"

    data = load_workbook_data(workbook_path)
    existing_tables = load_existing_tables(args.mysql_bin, args.database, args.user, args.password)

    with tempfile.NamedTemporaryFile("w", suffix=".sql", delete=False, encoding="utf-8") as tmp:
        sql_path = Path(tmp.name)
    try:
        write_sql(
            data,
            sql_path,
            map_refresh_path,
            args.batch_size,
            not args.preserve_upload_audit,
            existing_tables,
        )
        if not args.dry_run:
            run_mysql(args.mysql_bin, args.database, args.user, args.password, sql_path)
    finally:
        sql_path.unlink(missing_ok=True)

    mode = "Generated" if args.dry_run else "Imported"
    print(
        f"{mode} {data.stats['data_rows']} data rows, {len(data.compounds)} compounds, "
        f"{len(data.methods)} methods, {len(data.plants)} plants, "
        f"{data.stats['home_records']} home records, {data.stats['icd11_records']} ICD11 paths."
    )
    if data.stats["home_missing_doi"]:
        print(f"Warning: skipped {data.stats['home_missing_doi']} home rows without DOI.")


legacy_direct_main = main


def main() -> None:
    from import_wbe_workbook_via_api import main as api_import_main

    api_import_main()

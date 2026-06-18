#!/usr/bin/env python3
"""Import WBE workbook rows into the home_target_records fact table."""

from __future__ import annotations

import argparse
import os
import subprocess
import tempfile
from pathlib import Path

import openpyxl


DEFAULT_SUBCLASS = "默认"
DATA_SHEET = "数据表"
LITERATURE_SHEET = "文献基础信息"
INVALID_VALUES = {"", "NA", "N/A", "NULL", "NONE"}


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.upper() in INVALID_VALUES else text


def sql_value(value: object) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def row_target_group(target_category: str) -> str:
    return "drug" if target_category == "药物类" else "consumer"


def load_literature_dois(workbook: openpyxl.Workbook) -> dict[str, str]:
    sheet = workbook[LITERATURE_SHEET]
    header = [clean(value) for value in next(sheet.iter_rows(values_only=True))]
    column = {name: index for index, name in enumerate(header)}
    required = ["文献编号", "DOI"]
    missing = [name for name in required if name not in column]
    if missing:
        raise ValueError(f"{LITERATURE_SHEET} 缺少列: {', '.join(missing)}")

    dois: dict[str, str] = {}
    for row in sheet.iter_rows(values_only=True):
        literature_id = clean(row[column["文献编号"]])
        doi = clean(row[column["DOI"]])
        if literature_id and doi:
            dois[literature_id] = doi
    return dois


def build_records(workbook_path: Path) -> tuple[list[tuple[str, ...]], dict[str, int]]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    literature_dois = load_literature_dois(workbook)
    sheet = workbook[DATA_SHEET]
    header = [clean(value) for value in next(sheet.iter_rows(values_only=True))]
    column = {name: index for index, name in enumerate(header)}
    required = ["文献编号", "目标类别", "目标物质类别", "目标物质子类", "生物标记物名称"]
    missing = [name for name in required if name not in column]
    if missing:
        raise ValueError(f"{DATA_SHEET} 缺少列: {', '.join(missing)}")

    records: list[tuple[str, ...]] = []
    stats = {"skipped": 0, "missing_doi": 0}
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=2):
        literature_id = clean(row[column["文献编号"]])
        target_category = clean(row[column["目标类别"]])
        substance_category = clean(row[column["目标物质类别"]])
        substance_subclass = clean(row[column["目标物质子类"]]) or DEFAULT_SUBCLASS
        biomarker_name = clean(row[column["生物标记物名称"]])

        if literature_id == "文献编号" or target_category == "目标类别":
            stats["skipped"] += 1
            continue
        if not literature_id or not target_category or not substance_category or not biomarker_name:
            stats["skipped"] += 1
            continue

        doi = literature_dois.get(literature_id, "")
        if not doi:
            stats["missing_doi"] += 1
            continue

        records.append(
            (
                literature_id,
                doi,
                target_category,
                row_target_group(target_category),
                substance_category,
                substance_subclass,
                biomarker_name,
                DATA_SHEET,
                str(row_number),
            )
        )

    return records, stats


def write_import_sql(records: list[tuple[str, ...]], output_path: Path, batch_size: int) -> None:
    columns = (
        "literature_id",
        "doi",
        "target_category",
        "target_group",
        "substance_category",
        "substance_subclass",
        "biomarker_name",
        "source_sheet",
        "source_row_number",
    )
    with output_path.open("w", encoding="utf-8") as handle:
        handle.write("SET NAMES utf8mb4;\n")
        handle.write("TRUNCATE TABLE home_target_records;\n")
        for start in range(0, len(records), batch_size):
            chunk = records[start : start + batch_size]
            handle.write(f"INSERT INTO home_target_records ({', '.join(columns)}) VALUES\n")
            values = []
            for record in chunk:
                values.append("(" + ", ".join(sql_value(value) for value in record) + ")")
            handle.write(",\n".join(values))
            handle.write(";\n")


def run_mysql(mysql_bin: str, database: str, user: str, password: str, sql_path: Path) -> None:
    command = [mysql_bin, f"-u{user}", database]
    env = os.environ.copy()
    if password:
        env["MYSQL_PWD"] = password
    with sql_path.open("rb") as sql_file:
        subprocess.run(command, stdin=sql_file, env=env, check=True)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    backend_root = Path(__file__).resolve().parents[1]
    workspace_root = backend_root.parent
    parser.add_argument("--workbook", default=str(workspace_root / "WBE汇总表.xlsx"))
    parser.add_argument("--database", default="wbe")
    parser.add_argument("--user", default="root")
    parser.add_argument("--password", default="")
    parser.add_argument("--mysql-bin", default="mysql")
    parser.add_argument(
        "--schema",
        default=str(backend_root / "src/main/resources/db/home_target_records_schema.sql"),
    )
    parser.add_argument("--batch-size", type=int, default=500)
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    records, stats = build_records(workbook_path)
    if not records:
        raise RuntimeError("没有可导入的首页目标物质记录")

    schema_path = Path(args.schema).expanduser().resolve()
    run_mysql(args.mysql_bin, args.database, args.user, args.password, schema_path)
    with tempfile.NamedTemporaryFile("w", suffix=".sql", delete=False, encoding="utf-8") as tmp:
        import_sql = Path(tmp.name)
    try:
        write_import_sql(records, import_sql, args.batch_size)
        run_mysql(args.mysql_bin, args.database, args.user, args.password, import_sql)
    finally:
        import_sql.unlink(missing_ok=True)

    print(
        f"Imported {len(records)} records into home_target_records "
        f"(skipped={stats['skipped']}, missing_doi={stats['missing_doi']})."
    )


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Validate and match WBE workbook rows to the point-association sheet."""

from __future__ import annotations

import argparse
import json
import unicodedata
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable

import openpyxl


DATA_SHEET = "数据表"
SITE_SHEET = "点位关联表"
SITE_HEADERS = [
    "文献编号",
    "DOI",
    "国家",
    "省/州",
    "市",
    "原始污水厂名称",
    "规范污水厂名称",
    "reported_site_key",
    "confirmed_site_id",
    "是否计入点位数",
    "点位说明",
    "同一污水厂确认依据",
]
MISSING_TOKENS = {
    "",
    "-",
    "/",
    "na",
    "n/a",
    "n.a.",
    "nan",
    "none",
    "null",
    "未报道",
    "未报告",
    "未说明",
}

STATUS_EXACT = "精确关联"
STATUS_POSITION_FALLBACK = "位置字段回退匹配"
STATUS_MULTI = "一条记录关联多个点位"
STATUS_EXCLUDED = "关联记录不计数"
STATUS_UNMATCHED_COUNTRY = "未匹配国家"
STATUS_UNMATCHED = "未匹配"


def text(value: object) -> str:
    if value is None:
        return ""
    normalized = unicodedata.normalize("NFKC", str(value)).strip()
    return "" if normalized.lower() in MISSING_TOKENS else normalized


def normalized(value: object) -> str:
    return " ".join(text(value).lower().split())


def location_key(value: object) -> str:
    decomposed = unicodedata.normalize("NFKD", text(value))
    without_marks = "".join(char for char in decomposed if not unicodedata.combining(char))
    compact = "".join(char for char in without_marks.lower().replace("ı", "i") if char.isalnum())
    return {"tianjing": "tianjin"}.get(compact, compact)


def location_matches(site_value: object, record_value: object) -> bool:
    record_key = location_key(record_value)
    site_keys = {location_key(part) for part in text(site_value).split("/")}
    site_keys.discard("")
    if not site_keys or not record_key:
        return False
    return record_key in site_keys


def included(value: object) -> bool:
    return normalized(value) in {"是", "yes", "y", "true", "1"}


@dataclass(frozen=True)
class SiteRow:
    excel_row: int
    literature_code: str
    doi: str
    country: str
    province: str
    city: str
    raw_name: str
    canonical_name: str
    reported_site_key: str
    confirmed_site_id: str
    include_in_point_count: bool
    site_note: str
    confirmation_evidence: str

    @property
    def effective_site_key(self) -> str:
        return f"reported:{self.reported_site_key}"


@dataclass(frozen=True)
class RecordMatch:
    excel_row: int
    internal_record_key: str
    status: str
    reported_site_keys: tuple[str, ...]
    effective_site_keys: tuple[str, ...]


def header_map(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {text(cell.value): index for index, cell in enumerate(sheet[1]) if text(cell.value)}


def require_headers(sheet: openpyxl.worksheet.worksheet.Worksheet, required: Iterable[str]) -> dict[str, int]:
    headers = header_map(sheet)
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"{sheet.title} 缺少字段: {', '.join(missing)}")
    return headers


def row_values(row: tuple[openpyxl.cell.cell.Cell, ...], headers: dict[str, int]) -> dict[str, object]:
    return {name: row[index].value for name, index in headers.items() if index < len(row)}


def load_sites(sheet: openpyxl.worksheet.worksheet.Worksheet) -> tuple[list[SiteRow], list[str]]:
    headers = require_headers(sheet, SITE_HEADERS)
    sites: list[SiteRow] = []
    warnings: list[str] = []
    for excel_row, cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = row_values(cells, headers)
        if not any(text(values.get(name)) for name in SITE_HEADERS):
            continue
        site = SiteRow(
            excel_row=excel_row,
            literature_code=text(values["文献编号"]),
            doi=text(values["DOI"]),
            country=text(values["国家"]),
            province=text(values["省/州"]),
            city=text(values["市"]),
            raw_name=text(values["原始污水厂名称"]),
            canonical_name=text(values["规范污水厂名称"]),
            reported_site_key=text(values["reported_site_key"]),
            confirmed_site_id=text(values["confirmed_site_id"]),
            include_in_point_count=included(values["是否计入点位数"]),
            site_note=text(values["点位说明"]),
            confirmation_evidence=text(values["同一污水厂确认依据"]),
        )
        if site.include_in_point_count and not site.reported_site_key:
            raise ValueError(f"{SITE_SHEET} 第 {excel_row} 行计入点位数但 reported_site_key 为空")
        if not site.include_in_point_count and not site.site_note:
            warnings.append(f"{SITE_SHEET} 第 {excel_row} 行排除但缺少点位说明")
        if site.confirmed_site_id and not site.confirmation_evidence:
            warnings.append(f"{SITE_SHEET} 第 {excel_row} 行有 confirmed_site_id 但缺少确认依据")
        sites.append(site)

    duplicates = [
        key
        for key, count in Counter(
            (normalized(site.literature_code), site.reported_site_key) for site in sites
        ).items()
        if key[1] and count > 1
    ]
    if duplicates:
        raise ValueError(f"同一文献内 reported_site_key 重复: {len(duplicates)} 组")
    return sites, warnings


def narrow_by_location(candidates: list[SiteRow], province: str, city: str) -> list[SiteRow]:
    narrowed = candidates
    if city:
        narrowed = [site for site in narrowed if location_matches(site.city, city)]
    if province:
        narrowed = [site for site in narrowed if location_matches(site.province, province)]
    return narrowed


def match_record(
    excel_row: int,
    values: dict[str, object],
    sites_by_literature: dict[str, list[SiteRow]],
    batch_id: str,
) -> RecordMatch:
    literature_code = text(values.get("文献编号"))
    country = text(values.get("污水厂位置_国"))
    province = text(values.get("污水厂位置_省"))
    city = text(values.get("污水厂位置_市"))
    plant_name = text(values.get("污水厂名称"))
    internal_key = f"{batch_id}:{excel_row}"
    literature_candidates = sites_by_literature.get(normalized(literature_code), [])
    if not country:
        return RecordMatch(excel_row, internal_key, STATUS_UNMATCHED_COUNTRY, (), ())
    country_candidates = [
        site for site in literature_candidates if normalized(site.country) == normalized(country)
    ]
    if not country_candidates:
        return RecordMatch(excel_row, internal_key, STATUS_UNMATCHED_COUNTRY, (), ())
    location_candidates = narrow_by_location(country_candidates, province, city)
    if not location_candidates:
        return RecordMatch(excel_row, internal_key, STATUS_UNMATCHED, (), ())

    named_candidates: list[SiteRow] = []
    if plant_name:
        normalized_name = normalized(plant_name)
        named_candidates = [
            site
            for site in location_candidates
            if normalized_name in {normalized(site.raw_name), normalized(site.canonical_name)}
        ]
    else:
        counted_location_candidates = [site for site in location_candidates if site.include_in_point_count]
        if len(counted_location_candidates) == 1:
            named_candidates = counted_location_candidates
    chosen = named_candidates or location_candidates
    included_sites = [site for site in chosen if site.include_in_point_count]
    if not included_sites:
        return RecordMatch(
            excel_row,
            internal_key,
            STATUS_EXCLUDED,
            tuple(site.reported_site_key for site in chosen if site.reported_site_key),
            (),
        )
    if len(included_sites) > 1:
        status = STATUS_MULTI
    elif named_candidates:
        status = STATUS_EXACT
    else:
        status = STATUS_POSITION_FALLBACK
    return RecordMatch(
        excel_row,
        internal_key,
        status,
        tuple(site.reported_site_key for site in included_sites),
        tuple(site.effective_site_key for site in included_sites),
    )


def analyze_workbook(path: Path, batch_id: str = "workbook") -> dict[str, object]:
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        if DATA_SHEET not in workbook.sheetnames or SITE_SHEET not in workbook.sheetnames:
            raise ValueError(f"工作簿必须包含 {DATA_SHEET} 和 {SITE_SHEET}")
        site_rows, warnings = load_sites(workbook[SITE_SHEET])
        data_sheet = workbook[DATA_SHEET]
        data_headers = require_headers(
            data_sheet,
            ["文献编号", "污水厂名称", "污水厂位置_国", "污水厂位置_省", "污水厂位置_市"],
        )
        sites_by_literature: dict[str, list[SiteRow]] = defaultdict(list)
        for site in site_rows:
            sites_by_literature[normalized(site.literature_code)].append(site)

        matches: list[RecordMatch] = []
        record_values_by_row: dict[int, dict[str, object]] = {}
        for excel_row, cells in enumerate(data_sheet.iter_rows(min_row=2), start=2):
            values = row_values(cells, data_headers)
            if not any(text(value) for value in values.values()):
                continue
            record_values_by_row[excel_row] = values
            matches.append(match_record(excel_row, values, sites_by_literature, batch_id))

        included_sites = [site for site in site_rows if site.include_in_point_count]
        confirmed_groups = defaultdict(list)
        for site in site_rows:
            if site.confirmed_site_id:
                confirmed_groups[site.confirmed_site_id].append(site)
        mapped_site_keys = {
            key
            for match in matches
            for key in match.reported_site_keys
            if match.status not in {STATUS_EXCLUDED, STATUS_UNMATCHED, STATUS_UNMATCHED_COUNTRY}
        }
        site_by_key = {site.reported_site_key: site for site in site_rows}
        fallback_name_pairs: Counter[tuple[str, str, str]] = Counter()
        unmatched_records: list[dict[str, object]] = []
        excluded_records: list[dict[str, object]] = []
        excluded_site_key_counts: Counter[str] = Counter()
        for match in matches:
            if match.status == STATUS_UNMATCHED:
                values = record_values_by_row[match.excel_row]
                unmatched_records.append(
                    {
                        "excel_row": match.excel_row,
                        "literature_code": text(values.get("文献编号")),
                        "country": text(values.get("污水厂位置_国")),
                        "province": text(values.get("污水厂位置_省")),
                        "city": text(values.get("污水厂位置_市")),
                        "plant_name": text(values.get("污水厂名称")),
                    }
                )
            if match.status == STATUS_EXCLUDED:
                excluded_site_key_counts.update(match.reported_site_keys)
                values = record_values_by_row[match.excel_row]
                excluded_records.append(
                    {
                        "excel_row": match.excel_row,
                        "literature_code": text(values.get("文献编号")),
                        "country": text(values.get("污水厂位置_国")),
                        "province": text(values.get("污水厂位置_省")),
                        "city": text(values.get("污水厂位置_市")),
                        "plant_name": text(values.get("污水厂名称")),
                        "candidate_site_keys": list(match.reported_site_keys),
                    }
                )
            if match.status != STATUS_POSITION_FALLBACK or len(match.reported_site_keys) != 1:
                continue
            values = record_values_by_row[match.excel_row]
            site = site_by_key.get(match.reported_site_keys[0])
            if site:
                fallback_name_pairs[
                    (text(values.get("污水厂名称")), site.raw_name, site.canonical_name)
                ] += 1
        unmatched_sites = [
            {
                "reported_site_key": site.reported_site_key,
                "literature_code": site.literature_code,
                "country": site.country,
                "province": site.province,
                "city": site.city,
                "raw_name": site.raw_name,
            }
            for site in included_sites
            if site.reported_site_key not in mapped_site_keys
        ]
        china_site_count = len(
            {
                site.effective_site_key
                for site in included_sites
                if normalized(site.country) in {"中国", "china"}
            }
        )
        return {
            "workbook": str(path),
            "merge_confirmed_cross_document_sites": False,
            "site_rows": len(site_rows),
            "included_sites": len(included_sites),
            "excluded_sites": len(site_rows) - len(included_sites),
            "duplicate_reported_site_keys_within_literature": 0,
            "confirmed_candidate_groups": len(confirmed_groups),
            "confirmed_candidate_rows": sum(len(rows) for rows in confirmed_groups.values()),
            "record_rows": len(matches),
            "match_status_counts": dict(Counter(match.status for match in matches)),
            "mapped_sites": len(mapped_site_keys),
            "unmapped_sites": len(unmatched_sites),
            "china_point_count": china_site_count,
            "warnings": warnings,
            "unmatched_records": unmatched_records,
            "excluded_record_examples": excluded_records[:20],
            "excluded_site_key_counts": dict(excluded_site_key_counts.most_common()),
            "top_position_fallback_name_pairs": [
                {
                    "data_plant_name": names[0],
                    "site_raw_name": names[1],
                    "site_canonical_name": names[2],
                    "record_count": count,
                }
                for names, count in fallback_name_pairs.most_common(40)
            ],
            "unmatched_sites": unmatched_sites,
        }
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--batch-id", default="workbook")
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    report = analyze_workbook(args.workbook.expanduser().resolve(), args.batch_id)
    payload = json.dumps(report, ensure_ascii=False, indent=2)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(payload + "\n", encoding="utf-8")
    print(payload)


if __name__ == "__main__":
    main()
